import os
import pandas as pd
from tkinter import Tk, filedialog, Button, Label
from tkinter import messagebox
from openpyxl import load_workbook

# Função para renomear as colunas
def renomear_colunas_csv(pasta):
    mapeamento = { 
        "nome": "NOME",
        "telefone1": "TELEFONE",
        "email1": "EMAIL",
        "data_cadastro": "DATA CADASTRO",
        "telefone2": "TELEFONE",
        "tipo": "TIPO DE AÇÃO",
        "responsavel": "RESPONSÁVEL",
        "cnpj": "CPF CNPJ",
        "estado_civil": "ESTADO CIVIL",
        "cpf": "CPF CNPJ",
        "rg": "RG",
        "nascimento": "DATA DE NASCIMENTO",
        "profissao": "PROFISSÃO",
        "nacionalidade": "NACIONALIDADE",
        "logradouro": "ENDEREÇO",
        "bairro": "BAIRRO",
        "cep": "CEP",
        "cidade": "CIDADE",
        "telefone3": "TELEFONE",
        "email2": "EMAIL",
        "cliente": "NOME DO CLIENTE",
        "razao_social": "NOME DO CLIENTE",
        "telefone_comercial": "TELEFONE",
        "pis": "PIS PASEP",
        "nome_pai": "NOME DO PAI",
        "nome_mae": "NOME DA MÃE",
        "cpf_cnpj": "CPF CNPJ",
        "estado": "ESTADO",
        "fase": "FASE PROCESSUAL",
        "numero_processo": "NÚMERO DO PROCESSO",
        "pasta": "PASTA",
        "tipo_acao": "TIPO DE AÇÃO",
        "valor_causa": "EXPECTATIVA/VALOR DA CAUSA",
        "origem": "ORIGEM DO CLIENTE",
        "data_contratacao": "DATA DE CONTRATAÇÃO",
        "numero_vara": "VARA",
        "email": "EMAIL",
        "grupo_processo": "GRUPO DE AÇÃO",
        "nome_fantasia": "NOME DO CLIENTE",
    }

    # Processa cada arquivo CSV na pasta
    for arquivo in os.listdir(pasta):
        if arquivo.endswith(".csv"):
            caminho_arquivo = os.path.join(pasta, arquivo)
            try:
                # Lê o arquivo CSV
                df = pd.read_csv(caminho_arquivo, delimiter=';', encoding='ISO-8859-1')
                colunas_renomeadas = {col: mapeamento[col] for col in df.columns if col in mapeamento}
                df.rename(columns=colunas_renomeadas, inplace=True)
                df.to_csv(caminho_arquivo, index=False, sep=';', encoding='ISO-8859-1')
            except Exception as e:
                print(f"Erro ao processar o arquivo '{arquivo}': {e}")

# Função para inserir dados nas planilhas Excel
def inserir_dados_excel(pasta):
    caminho_arquivo_processos = "Advbox/PROCESSOS.xlsx"
    caminho_arquivo_clientes = "Advbox/CLIENTES.xlsx"

    try:
        wb_processos = load_workbook(caminho_arquivo_processos)
        sheet_processos = wb_processos['Página1']
        xlsx_df_processos = pd.read_excel(caminho_arquivo_processos, sheet_name='Página1')
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar o arquivo 'PROCESSOS': {e}")
        return

    for arquivo in os.listdir(pasta):
        if arquivo.endswith(".csv"):
            caminho_csv = os.path.join(pasta, arquivo)
            try:
                csv_df = pd.read_csv(caminho_csv, delimiter=';', encoding='ISO-8859-1')
                colunas_comum = list(set(csv_df.columns).intersection(set(xlsx_df_processos.columns)))
                if colunas_comum:
                    for coluna in colunas_comum:
                        col_idx = xlsx_df_processos.columns.get_loc(coluna) + 1
                        for i, linha in csv_df.iterrows():
                            sheet_processos.cell(row=i + 2, column=col_idx, value=linha[coluna])
                    wb_processos.save(caminho_arquivo_processos)
            except Exception as e:
                print(f"Erro ao processar o arquivo CSV '{arquivo}': {e}")

    try:
        wb_clientes = load_workbook(caminho_arquivo_clientes)
        sheet_clientes = wb_clientes['Página1']
        xlsx_df_clientes = pd.read_excel(caminho_arquivo_clientes, sheet_name='Página1')
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao carregar o arquivo 'CLIENTES': {e}")
        return

    for arquivo in os.listdir(pasta):
        if arquivo.endswith(".csv"):
            caminho_csv = os.path.join(pasta, arquivo)
            try:
                csv_df = pd.read_csv(caminho_csv, delimiter=';', encoding='ISO-8859-1')
                colunas_comum = list(set(csv_df.columns).intersection(set(xlsx_df_clientes.columns)))
                if colunas_comum:
                    for i, linha in csv_df.iterrows():
                        if pd.notna(linha['NOME']) and linha['NOME'] != '':
                            for coluna in colunas_comum:
                                col_idx = xlsx_df_clientes.columns.get_loc(coluna) + 1
                                sheet_clientes.cell(row=i + 2, column=col_idx, value=linha[coluna])
                    wb_clientes.save(caminho_arquivo_clientes)
            except Exception as e:
                print(f"Erro ao processar o arquivo CSV '{arquivo}': {e}")

# Função para juntar as planilhas em um novo arquivo Excel
def juntar_planilhas():
    try:
        clientes_df = pd.read_excel("Advbox/CLIENTES.xlsx")
        processos_df = pd.read_excel("Advbox/PROCESSOS.xlsx")
        output_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if output_path:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                clientes_df.to_excel(writer, sheet_name='CLIENTES', index=False)
                processos_df.to_excel(writer, sheet_name='PROCESSOS', index=False)
            messagebox.showinfo("Sucesso", f'Planilhas foram juntadas com sucesso em: {output_path}')
        else:
            messagebox.showwarning("Aviso", "Nenhum caminho de saída foi selecionado.")
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao juntar as planilhas: {e}")

# Função principal da interface
def main():
    # Criação da janela
    root = Tk()
    root.title("Juntar Planilhas e Processar CSVs")

    # Função de selecionar a pasta
    def selecionar_pasta():
        pasta = filedialog.askdirectory(title="Selecione a pasta contendo os arquivos CSV")
        if pasta:
            # Processamento dos arquivos
            renomear_colunas_csv(pasta)
            inserir_dados_excel(pasta)
            messagebox.showinfo("Sucesso", "Processamento concluído com sucesso!")

    # Criar os botões da interface
    Button(root, text="Selecionar Pasta e Processar Arquivos", command=selecionar_pasta).pack(pady=10)
    Button(root, text="Juntar Planilhas", command=juntar_planilhas).pack(pady=10)

    # Adicionar um rótulo informativo
    Label(root, text="Selecione a pasta contendo os arquivos CSV e clique em 'Processar'").pack(pady=20)

    # Iniciar a interface
    root.mainloop()

if __name__ == "__main__":
    main()
