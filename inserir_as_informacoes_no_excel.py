import os
import pandas as pd
from openpyxl import load_workbook

caminho_pasta_dados = "Dados"
caminho_arquivo_processos = "Advbox/PROCESSOS.xlsx"
caminho_arquivo_clientes = "Advbox/CLIENTES.xlsx"

try:
    wb_processos = load_workbook(caminho_arquivo_processos)
    sheet_processos = wb_processos['Página1']
    xlsx_df_processos = pd.read_excel(caminho_arquivo_processos, sheet_name='Página1')
    print(f"Planilhas encontradas no Excel 'PROCESSOS': {wb_processos.sheetnames}")
except Exception as e:
    print(f"Erro ao carregar o arquivo Excel 'PROCESSOS': {e}")
    exit()

for arquivo in os.listdir(caminho_pasta_dados):
    if arquivo.endswith(".csv"):
        caminho_csv = os.path.join(caminho_pasta_dados, arquivo)

        try:
            csv_df = pd.read_csv(caminho_csv, delimiter=';', encoding='ISO-8859-1')

            colunas_comum = list(set(csv_df.columns).intersection(set(xlsx_df_processos.columns)))

            if colunas_comum:
                for coluna in colunas_comum:
                    col_idx = xlsx_df_processos.columns.get_loc(coluna) + 1  # Índice da coluna no Excel (1-based)

                    for i, linha in csv_df.iterrows():
                        sheet_processos.cell(row=i + 2, column=col_idx, value=linha[coluna])  # Começar a partir da linha 2 (evitar sobrescrever os nomes das colunas)

                    print(f"Coluna '{coluna}' do arquivo '{arquivo}' foi adicionada à planilha 'PROCESSOS'.")

            wb_processos.save(caminho_arquivo_processos)

        except Exception as e:
            print(f"Erro ao processar o arquivo CSV '{arquivo}': {e}")

try:
    wb_clientes = load_workbook(caminho_arquivo_clientes)
    sheet_clientes = wb_clientes['Página1']
    xlsx_df_clientes = pd.read_excel(caminho_arquivo_clientes, sheet_name='Página1')
    print(f"Planilhas encontradas no Excel 'CLIENTES': {wb_clientes.sheetnames}")
except Exception as e:
    print(f"Erro ao carregar o arquivo Excel 'CLIENTES': {e}")
    exit()

for arquivo in os.listdir(caminho_pasta_dados):
    if arquivo.endswith(".csv"):
        caminho_csv = os.path.join(caminho_pasta_dados, arquivo)

        try:
            csv_df = pd.read_csv(caminho_csv, delimiter=';', encoding='ISO-8859-1')

            colunas_comum = list(set(csv_df.columns).intersection(set(xlsx_df_clientes.columns)))

            if colunas_comum:
                for i, linha in csv_df.iterrows():
                    # Verificar se a coluna "NOME" está preenchida
                    if pd.notna(linha['NOME']) and linha['NOME'] != '':
                        # Adicionar as linhas apenas se a coluna "NOME" estiver preenchida
                        for coluna in colunas_comum:
                            col_idx = xlsx_df_clientes.columns.get_loc(coluna) + 1  # Índice da coluna no Excel (1-based)
                            sheet_clientes.cell(row=i + 2, column=col_idx, value=linha[coluna])  # Começar a partir da linha 2 (evitar sobrescrever os nomes das colunas)

                        print(f"Coluna '{linha['NOME']}' do arquivo '{arquivo}' foi adicionada à planilha 'CLIENTES'.")

            wb_clientes.save(caminho_arquivo_clientes)

        except Exception as e:
            print(f"Erro ao processar o arquivo CSV '{arquivo}': {e}")
